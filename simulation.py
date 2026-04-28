#!/usr/bin/env python
# coding: utf-8
import matplotlib
matplotlib.use('Agg')  # save plots to files instead of opening windows

# Patch plt.show globally so every module auto-saves instead of trying to open a window
import matplotlib.pyplot as _plt
_fig_counter = [0]
def _patched_show(*args, **kwargs):
    _fig_counter[0] += 1
    path = f'output_figure_{_fig_counter[0]:02d}.png'
    _plt.savefig(path, dpi=150, bbox_inches='tight')
    _plt.close()
    print(f'  -> saved {path}')
_plt.show = _patched_show

def _savefig():
    _patched_show()

# # Booster Impact Tracker — Velocity Frame Edition (v2.1)
# 
# Run each cell with **Shift+Enter**.  
# Change any parameter and re-run from that cell down (`Run → Run All Below`).
# 
# ### Velocity-frame recap
# | Symbol | Name | Description |
# |--------|------|-------------|
# | **γV** | Flight-path angle | Angle of velocity vector above horizontal (deg). Positive = climbing. |
# | **γH** | Heading azimuth | Direction of horizontal velocity (deg). 0 = North, 90 = East. |
# | **ê_t** | Tangential axis | Along velocity vector |
# | **ê_nV** | Pitch-normal axis | ⊥ to velocity, upward in pitch plane — lift & pitch commands act here |
# | **ê_nH** | Yaw-normal axis | Lateral direction — yaw commands act here |
# 
# ### Changelog
# | Version | Change |
# |---------|--------|
# | v2.1 | IACPN hit angle constraints (`hit_gamma_v`, `hit_gamma_h`, `hit_angle_range`); guidance active from launch (yaw-only during burn, full PN+IACPN post-burnout) |
# | v2.0 | Velocity-frame formulation, gravity turn, lift, 3-D heading, PN guidance |

# ## 1 — Imports

# In[1]:


from sim.simulate import simulate
from sim.physics import Params
from plot import plot_results
from dataclasses import replace
import matplotlib.pyplot as plt
# get_ipython().run_line_magic('matplotlib', 'inline')
print('Imports OK')


# ## 2 — Parameters
# 
# > **Gravity turn tip:** use `launch_angle` ≥ 85° with `grav_turn=True`.  
# > At 75° the rapid pitch-over at low speed keeps max altitude very low.  
# > With `grav_turn=False` any angle works (thrust held at fixed body angle).

# In[3]:


params = Params(
    # ── Booster ──────────────────────────────────────
    m_pay        = 50,     # kg  payload mass
    m_prop       = 495,    # kg  propellant mass
    m_str        = 50,     # kg  structural dry mass
    # ── Motor ────────────────────────────────────────
    isp          = 260,    # s
    thrust_kn    = 31,     # kN
    burn_max     = 20,     # s
    # ── Aerodynamics ─────────────────────────────────
    cd           = 0.4,    # drag coeff – powered
    cd_fall      = 1.2,    # drag coeff – tumbling post-burnout
    cl           = 0.0,    # lift coefficient  (0 = no lift)
    cd_ctrl      = 0.3,    # actuator drag (PN guidance)
    diam         = 0.5,    # m  reference diameter
    atm          = 'isa',  # 'isa' | 'exp' | 'none'
    # ── Launch angles ─────────────────────────────────
    launch_angle   = 45,   # deg  initial γV (elevation)
    launch_azimuth = 45,   # deg  aimed NE toward (10, 10) km target
    # ── Velocity-frame dynamics ───────────────────────
    grav_turn       = True,  # True → thrust ∥ velocity; False → fixed body angle
    grav_turn_v_min = 30.0,  # m/s  minimum speed before gravity turn activates
    a_cmd_nv        = 0.0,   # m/s²  pitch-normal accel command (ê_nV)
    a_cmd_nh        = 0.0,   # m/s²  yaw-normal accel command (ê_nH)
    # ── PN/IACPN guidance target ──────────────────────
    x_target     = 10,     # km  North
    y_target     = 10,     # km  East
    z_target     = 0,      # km  altitude
    a_lat_max    = 20.0,    # g   lateral accel limit
    # ── Impact angle constraints (optional) ──────────
    hit_gamma_v     = None,  # deg  desired impact flight-path angle (e.g. -60); None = unconstrained
    hit_gamma_h     = None,  # deg  desired impact azimuth (e.g. 45); None = unconstrained
    hit_angle_range = None,  # km   range from target at which angle law activates; None = always
    # ── Velocity-dependent acceleration limit (optional) ─────────────────
    # List of (speed m/s, a_lat_max g) breakpoints — linearly interpolated.
    # Overrides the constant a_lat_max above when set.
    # Example: low limit at low speed, higher limit at high speed.
    # a_lat_max_table = [(0, 5.0), (200, 10.0), (500, 20.0)],
)


# ## 3 — Run simulation

# In[4]:


result = simulate(params)

print('=== Summary ===')
for k, v in result['summary'].items():
    print(f'  {k:<30} {v}')


# ### Export state vector to Excel
import openpyxl

def export_state_vector(result, path='state_vector.xlsx'):
    s = result['series']
    wb = openpyxl.Workbook()

    # ── Sheet 1: State vector time series ──────────────────────────────────
    ws = wb.active
    ws.title = 'State Vector'
    columns = [
        ('t (s)',          't'),
        ('x N (km)',       'x'),
        ('y E (km)',       'y'),
        ('h alt (km)',     'h'),
        ('v (m/s)',        'v'),
        ('gamma_v (deg)',  'gamma_v'),
        ('gamma_h (deg)',  'gamma_h'),
        ('thrust (kN)',    'thr'),
        ('drag_aero (kN)', 'drag_aero'),
        ('drag_ctrl (kN)', 'drag_ctrl'),
        ('lift (kN)',      'lift'),
        ('acc (g)',        'acc'),
        ('aLat (g)',        'aLat'),
        ('a_nV (g)',        'a_nV'),
        ('a_nH (g)',        'a_nH'),
        ('mass (kg)',       'mass'),
        ('a_lat_max (g)',   'a_lat_max'),
    ]
    ws.append([col for col, _ in columns])
    for i in range(len(s['t'])):
        ws.append([s[key][i] for _, key in columns])

    # ── Sheet 2: Summary ───────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.append(['Parameter', 'Value'])
    for k, v in result['summary'].items():
        ws2.append([k, v])

    wb.save(path)
    print(f'  -> saved {path}')

export_state_vector(result)


def export_grid_excel(runs, prefix='grid'):
    """Write one Excel file per run: state vector + summary sheets."""
    # Column definitions — only written if the key exists in the series
    all_columns = [
        ('t (s)',          't'),
        ('x N (km)',       'x'),
        ('y E (km)',       'y'),
        ('h alt (km)',     'h'),
        ('v (m/s)',        'v'),
        ('gamma_v (deg)',  'gamma_v'),
        ('gamma_h (deg)',  'gamma_h'),
        ('thrust (kN)',    'thr'),
        ('drag_aero (kN)', 'drag_aero'),
        ('drag_ctrl (kN)', 'drag_ctrl'),
        ('lift (kN)',      'lift'),
        ('acc (g)',        'acc'),
        ('aLat (g)',       'aLat'),
        ('a_nV (g)',       'a_nV'),
        ('a_nH (g)',       'a_nH'),
        ('mass (kg)',      'mass'),
        ('a_lat_max (g)',  'a_lat_max'),
    ]

    for run in runs:
        s  = run['series']
        sm = run['summary']
        x, y, z = run['hit_point']
        gv, gh  = run['hit_angles']

        # Build safe filename
        angle_tag = ''
        if gv is not None: angle_tag += f'_gv{int(gv)}'
        if gh is not None: angle_tag += f'_gh{int(gh)}'
        fname = f'{prefix}_run{run["index"]+1:02d}_x{x}_y{y}{angle_tag}.xlsx'

        wb = openpyxl.Workbook()

        # ── State vector sheet ─────────────────────────────────────────────
        ws = wb.active
        ws.title = 'State Vector'
        columns = [(header, key) for header, key in all_columns if key in s]
        ws.append([header for header, _ in columns])
        for i in range(len(s['t'])):
            ws.append([s[key][i] for _, key in columns])

        # ── Summary sheet ──────────────────────────────────────────────────
        ws2 = wb.create_sheet('Summary')
        ws2.append(['Parameter', 'Value'])
        ws2.append(['target_x_km', x])
        ws2.append(['target_y_km', y])
        ws2.append(['target_z_km', z])
        ws2.append(['hit_gamma_v_deg', gv])
        ws2.append(['hit_gamma_h_deg', gh])
        for k, v in sm.items():
            ws2.append([k, v])

        wb.save(fname)
        print(f'  -> saved {fname}')


# ## 4 — Plots (7-panel dashboard)

# In[5]:


plot_results(result, params)


# ## 5 — Inspect time-series
# All arrays live in `result['series']`.

# In[6]:


s = result['series']
print('Available keys:', list(s.keys()))
print(f"Samples: {len(s['t'])}  |  Flight time: {s['t'][-1]} s")


# In[10]:


i = 10   # <-- change to inspect any timestep

print(f"t = {s['t'][i]} s")
print(f"  Position :  h={s['h'][i]} km   x={s['x'][i]} km   y={s['y'][i]} km")
print(f"  Velocity :  v={s['v'][i]} m/s")
print(f"  Frame    :  γV={s['gamma_v'][i]}°   γH={s['gamma_h'][i]}°")
print(f"  Forces   :  thrust={s['thr'][i]} kN   drag={s['drag_aero'][i]} kN   lift={s['lift'][i]} kN")
print(f"  ZEM      :  |aLat|={s['aLat'][i]} g   (nV={s['a_nV'][i]} g, nH={s['a_nH'][i]} g)")
print(f"  Accel    :  {s['acc'][i]} g")


# ## 6 — Flight-path angle profile (γV and γH)
# Shows how the velocity-frame angles evolve over the full flight.

# In[7]:


fig, ax = plt.subplots(figsize=(10, 4))
ax.plot(s['t'], s['gamma_v'], color='#378ADD', linewidth=2, label='γV — flight-path angle')
ax.plot(s['t'], s['gamma_h'], color='#7F77DD', linewidth=1.5, linestyle='--', label='γH — heading azimuth')
ax.axhline(0, color='#999', linewidth=0.8, linestyle=':')
ax.set_xlabel('Time (s)')
ax.set_ylabel('degrees')
ax.set_title('Velocity-frame angles vs time')
ax.legend()
ax.grid(alpha=0.3)
plt.tight_layout()
_savefig()


# ## 7 — Gravity turn vs fixed-angle comparison
# Runs the same vehicle twice and overlays the altitude and γV profiles.

# In[8]:


r_gt  = simulate(replace(params, grav_turn=True,  a_lat_max=0))  # no ZEM, pure ballistic
r_fix = simulate(replace(params, grav_turn=False, a_lat_max=0))

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4))

# Altitude
ax1.plot(r_gt['series']['t'],  r_gt['series']['h'],  color='#378ADD', label='Gravity turn')
ax1.plot(r_fix['series']['t'], r_fix['series']['h'], color='#D85A30', linestyle='--', label='Fixed angle')
ax1.set_xlabel('Time (s)'); ax1.set_ylabel('Altitude (km)')
ax1.set_title('Altitude comparison'); ax1.legend(); ax1.grid(alpha=0.3)

# γV
ax2.plot(r_gt['series']['t'],  r_gt['series']['gamma_v'],  color='#378ADD', label='Gravity turn')
ax2.plot(r_fix['series']['t'], r_fix['series']['gamma_v'], color='#D85A30', linestyle='--', label='Fixed angle')
ax2.axhline(0, color='#999', linewidth=0.8, linestyle=':')
ax2.set_xlabel('Time (s)'); ax2.set_ylabel('γV (deg)')
ax2.set_title('Flight-path angle γV'); ax2.legend(); ax2.grid(alpha=0.3)

plt.tight_layout(); _savefig()

for label, r in [('Gravity turn', r_gt), ('Fixed angle', r_fix)]:
    sm = r['summary']
    print(f"{label:<15}  max_alt={sm['max_altitude_km']:.1f} km  "
          f"burnout_gV={sm['burnout_gamma_v_deg']:.1f}°  "
          f"range={sm['impact_range_km']:.1f} km")


# ## 8 — Lift coefficient sweep
# Shows how increasing CL slows the gravity-turn pitch-over (higher γV at burnout → more altitude).

# In[ ]:


cl_values = [0.0, 0.2, 0.5, 1.0]
colors     = ['#378ADD', '#1D9E75', '#D85A30', '#7F77DD']

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4))

print(f"{'CL':>5}  {'max_alt':>10}  {'burnout_gV':>12}  {'max_lift':>10}  {'range':>10}")
for cl, col in zip(cl_values, colors):
    r  = simulate(replace(params, cl=cl, a_lat_max=0))
    sm = r['summary']
    sv = r['series']
    ax1.plot(sv['t'], sv['h'],       color=col, label=f'CL={cl}')
    ax2.plot(sv['t'], sv['gamma_v'], color=col, label=f'CL={cl}')
    print(f"{cl:>5}  {sm['max_altitude_km']:>10.2f}  "
          f"{sm['burnout_gamma_v_deg']:>12.1f}°  "
          f"{sm['max_lift_kn']:>10.3f} kN  "
          f"{sm['impact_range_km']:>10.1f} km")

ax1.set_xlabel('Time (s)'); ax1.set_ylabel('Altitude (km)')
ax1.set_title('Altitude vs CL'); ax1.legend(); ax1.grid(alpha=0.3)
ax2.axhline(0, color='#999', linewidth=0.8, linestyle=':')
ax2.set_xlabel('Time (s)'); ax2.set_ylabel('γV (deg)')
ax2.set_title('Flight-path angle γV vs CL'); ax2.legend(); ax2.grid(alpha=0.3)
plt.tight_layout(); _savefig()


# In[ ]:





# In[ ]:





# cl_values = [0.0, 0.2, 0.5, 1.0]
# colors     = ['#378ADD', '#1D9E75', '#D85A30', '#7F77DD']
# 
# fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4))
# 
# print(f"{'CL':>5}  {'max_alt':>10}  {'burnout_gV':>12}  {'max_lift':>10}  {'range':>10}")
# for cl, col in zip(cl_values, colors):
#     r  = simulate(replace(params, cl=cl, a_lat_max=0))
#     sm = r['summary']
#     sv = r['series']
#     ax1.plot(sv['t'], sv['h'],       color=col, label=f'CL={cl}')
#     ax2.plot(sv['t'], sv['gamma_v'], color=col, label=f'CL={cl}')
#     print(f"{cl:>5}  {sm['max_altitude_km']:>10.2f}  "
#           f"{sm['burnout_gamma_v_deg']:>12.1f}°  "
#           f"{sm['max_lift_kn']:>10.3f} kN  "
#           f"{sm['impact_range_km']:>10.1f} km")
# 
# ax1.set_xlabel('Time (s)'); ax1.set_ylabel('Altitude (km)')
# ax1.set_title('Altitude vs CL'); ax1.legend(); ax1.grid(alpha=0.3)
# ax2.axhline(0, color='#999', linewidth=0.8, linestyle=':')
# ax2.set_xlabel('Time (s)'); ax2.set_ylabel('γV (deg)')
# ax2.set_title('Flight-path angle γV vs CL'); ax2.legend(); ax2.grid(alpha=0.3)
# plt.tight_layout(); plt.show()

# ## 9 — Pitch acceleration command sweep (a_cmd_nv)
# `a_cmd_nv` acts along **ê_nV** (pitch-normal). Positive pulls the flight path upward, slowing the gravity-turn pitch-over.

# In[14]:


nv_values = [-2.0, 0.0, 2.0, 5.0]
colors     = ['#D85A30', '#378ADD', '#1D9E75', '#7F77DD']

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4))

print(f"{'a_nV':>6}  {'max_alt':>10}  {'burnout_gV':>12}  {'range':>10}")
for nv, col in zip(nv_values, colors):
    r  = simulate(replace(params, a_cmd_nv=nv, a_lat_max=0))
    sm = r['summary']
    sv = r['series']
    ax1.plot(sv['t'], sv['h'],       color=col, label=f'a_nV={nv:+.1f} m/s²')
    ax2.plot(sv['t'], sv['gamma_v'], color=col, label=f'a_nV={nv:+.1f} m/s²')
    print(f"{nv:>+6.1f}  {sm['max_altitude_km']:>10.2f}  "
          f"{sm['burnout_gamma_v_deg']:>12.1f}°  "
          f"{sm['impact_range_km']:>10.1f} km")

ax1.set_xlabel('Time (s)'); ax1.set_ylabel('Altitude (km)')
ax1.set_title('Altitude vs a_cmd_nV'); ax1.legend(fontsize=8); ax1.grid(alpha=0.3)
ax2.axhline(0, color='#999', linewidth=0.8, linestyle=':')
ax2.set_xlabel('Time (s)'); ax2.set_ylabel('γV (deg)')
ax2.set_title('Flight-path angle γV vs a_cmd_nV'); ax2.legend(fontsize=8); ax2.grid(alpha=0.3)
plt.tight_layout(); _savefig()


# ## 10 — Yaw command sweep (a_cmd_nh)
# `a_cmd_nh` acts along **ê_nH** (yaw-normal). It changes γH over time, diverting the booster laterally. The impact range uses 3-D distance √(x²+y²).

# In[15]:


import math
nh_values = [0.0, 1.0, 3.0, -2.0]
colors     = ['#378ADD', '#1D9E75', '#D85A30', '#7F77DD']

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 4))

print(f"{'a_nH':>6}  {'final_gH':>10}  {'impact_x':>10}  {'impact_y':>10}  {'range':>10}")
for nh, col in zip(nh_values, colors):
    r  = simulate(replace(params, a_cmd_nh=nh, a_lat_max=0))
    sm = r['summary']
    sv = r['series']
    # Ground track
    ax1.plot(sv['x'], sv['y'], color=col, label=f'a_nH={nh:+.1f} m/s²')
    ax2.plot(sv['t'], sv['gamma_h'], color=col, label=f'a_nH={nh:+.1f} m/s²')
    print(f"{nh:>+6.1f}  {sm['final_gamma_h_deg']:>10.1f}°  "
          f"{sm['impact_x_km']:>10.2f} km  "
          f"{sm['impact_y_km']:>10.2f} km  "
          f"{sm['impact_range_km']:>10.1f} km")

ax1.set_xlabel('North X (km)'); ax1.set_ylabel('East Y (km)')
ax1.set_title('Ground track vs a_cmd_nH'); ax1.legend(fontsize=8); ax1.grid(alpha=0.3)
ax2.set_xlabel('Time (s)'); ax2.set_ylabel('γH (deg)')
ax2.set_title('Heading azimuth γH vs a_cmd_nH'); ax2.legend(fontsize=8); ax2.grid(alpha=0.3)
plt.tight_layout(); _savefig()


# ## 11 — Launch angle sweep
# Classic trade: higher angle → more altitude, less range.

# In[16]:


angles = [30, 45, 60, 75]

print(f"{'angle':>6}  {'max_alt':>10}  {'burnout_gV':>12}  {'burnout_alt':>12}  {'range':>10}  {'t_flight':>10}")
for angle in angles:
    r  = simulate(replace(params, launch_angle=angle, a_lat_max=0))
    sm = r['summary']
    print(f"{angle:>6}°  {sm['max_altitude_km']:>10.2f}  "
          f"{sm['burnout_gamma_v_deg']:>12.1f}°  "
          f"{sm['burnout_altitude_km']:>12.2f} km  "
          f"{sm['impact_range_km']:>10.1f} km  "
          f"{sm['flight_time_s']:>10.1f} s")


# ## 12 — Run check vectors (pytest)

# In[17]:


import subprocess; subprocess.run(['python', '-m', 'pytest', 'tests/check_vectors.py', '-v'])


# ## 13 — Impact angle constraints (IACPN)
# 
# Guidance is **active from launch**:
# - **During burn** — yaw channel only (`a_nH`): heading is corrected toward the target while pitch is left to the gravity turn
# - **After burnout** — full PN + IACPN on both pitch (`a_nV`) and yaw (`a_nH`) channels
# 
# Two optional terminal angle constraints can be added on top of PN:
# 
# | Parameter | Description |
# |-----------|-------------|
# | `hit_gamma_v` | Desired impact flight-path angle γV (deg). e.g. `−60` = 60° below horizontal; `−90` = vertical dive. `None` = unconstrained. |
# | `hit_gamma_h` | Desired impact azimuth γH (deg). Same convention as `launch_azimuth`. `None` = unconstrained. |
# | `hit_angle_range` | Range from target (km) at which the IACPN angle correction activates. Outside = pure PN; inside = IACPN. `None` = always active. |
# 
# When inside the activation range, the IACPN adds a tgo-weighted correction on top of PN:
# 
# ```
# a_nV_angle = K · v · sin(hit_γV − γV_LOS) / tgo
# a_nH_angle = K · v · cos(γV) · sin(hit_γH − γH_LOS) / tgo
# ```
# 
# > **Note:** angle constraints trade miss distance for angle compliance.

# In[18]:


from dataclasses import replace

base = Params(
    m_pay=50, m_prop=495, m_str=50, isp=260, thrust_kn=31, burn_max=20,
    cd=0.4, cd_fall=1.2, cl=0.0, cd_ctrl=0.3, diam=0.5,
    launch_angle=45, launch_azimuth=45,
    x_target=10, y_target=10, z_target=0, a_lat_max=3.0,
)

cases = [
    ('Pure PN (no angles)',         dict(hit_gamma_v=None, hit_gamma_h=None, hit_angle_range=None)),
    ('IACPN, range gate 15 km',    dict(hit_gamma_v=-60,  hit_gamma_h=45,   hit_angle_range=15)),
    ('IACPN, range gate  5 km',    dict(hit_gamma_v=-60,  hit_gamma_h=45,   hit_angle_range=5)),
    ('IACPN, always active',       dict(hit_gamma_v=-60,  hit_gamma_h=45,   hit_angle_range=None)),
    ('Vertical only (γV=-80)',      dict(hit_gamma_v=-80,  hit_gamma_h=None, hit_angle_range=15)),
]

print(f"{'Case':<30}  {'miss_m':>8}  {'γV_impact':>10}  {'γH_impact':>10}")
print('-' * 65)
for label, kw in cases:
    r = simulate(replace(base, **kw))
    sm = r['summary']
    print(f"{label:<30}  {sm['miss_distance_m']:>8.1f}  "
          f"{sm['final_gamma_v_deg']:>9.1f}°  "
          f"{sm['final_gamma_h_deg']:>9.1f}°")


# ## 14 — IACPN angle constraint comparison
# Guidance is active from launch. This section compares pure PN against IACPN with different angle constraints and range gates.

# In[19]:


base = Params(
    m_pay=50, m_prop=495, m_str=50, isp=260, thrust_kn=31, burn_max=20,
    cd=0.4, cd_fall=1.2, cl=0.0, cd_ctrl=0.3, diam=0.5,
    launch_angle=45, launch_azimuth=45,
    x_target=10, y_target=10, z_target=0, a_lat_max=3.0,
)

cases = [
    ('Pure PN',                    dict(hit_gamma_v=None, hit_gamma_h=None, hit_angle_range=None)),
    ('IACPN γV=-60 gate=15km',    dict(hit_gamma_v=-60,  hit_gamma_h=None, hit_angle_range=15)),
    ('IACPN γV=-60 gate=5km',     dict(hit_gamma_v=-60,  hit_gamma_h=None, hit_angle_range=5)),
    ('IACPN γV=-60 γH=45 gate=15',dict(hit_gamma_v=-60,  hit_gamma_h=45,   hit_angle_range=15)),
    ('IACPN γV=-80 always',       dict(hit_gamma_v=-80,  hit_gamma_h=None, hit_angle_range=None)),
]

print(f"{'Case':<30}  {'miss_m':>8}  {'γV_impact':>10}  {'γH_impact':>10}")
print('-' * 65)
for label, kw in cases:
    r = simulate(replace(base, **kw))
    sm = r['summary']
    print(f"{label:<30}  {sm['miss_distance_m']:>8.1f}  "
          f"{sm['final_gamma_v_deg']:>9.1f}°  "
          f"{sm['final_gamma_h_deg']:>9.1f}°")

# ── Ground tracks ─────────────────────────────────────────────────────────────
colors = ['#378ADD', '#1D9E75', '#D85A30', '#7F77DD', '#E8A838']
fig, ax = plt.subplots(figsize=(7, 6))
for (label, kw), col in zip(cases, colors):
    r = simulate(replace(base, **kw))
    sv = r['series']
    ax.plot(sv['x'], sv['y'], color=col, label=label)
ax.plot(base.x_target, base.y_target, 'k*', markersize=14, label='Target')
ax.set_xlabel('North X (km)'); ax.set_ylabel('East Y (km)')
ax.set_title('Ground tracks — PN vs IACPN angle constraints')
ax.legend(fontsize=8); ax.grid(alpha=0.3)
plt.tight_layout(); _savefig()


# ## 15 — Grid runner
# 
# `run_grid()` sweeps a **Cartesian product** of hit points × hit angles. All other parameters come from `base_params`.
# 
# ```python
# run_grid(base_params, hit_points, hit_angles=None, hit_angle_range=None)
# ```
# 
# | Argument | Type | Description |
# |----------|------|-------------|
# | `base_params` | `Params` | Shared parameters for all runs |
# | `hit_points` | `[(x,y,z), ...]` km | Target locations to hit |
# | `hit_angles` | `[(γV,γH), ...]` deg or `None` | Angle constraints — cross-product with hit_points. Either element may be `None`. Omit entirely for no constraints. |
# | `hit_angle_range` | `float \| None` km | Range gate for angle law, applied to all constrained runs |
# 
# `plot_grid(runs)` shows a 4-panel dashboard: ground tracks, altitude profiles, γV profiles, miss distance bar chart + printed summary table.

# In[20]:


from grid_run import run_grid, plot_grid

base_grid = Params(
    m_pay=50, m_prop=100, m_str=50, isp=260, thrust_kn=31, burn_max=20,
    cd=0.4, cd_fall=0.4, cl=0.0, cd_ctrl=0.3, diam=0.5,
    launch_angle=45, launch_azimuth=45, a_lat_max=20.0,
)

# Hit points at very different ranges AND azimuths — maximally distinct paths
hit_points = [
    ( 5,  0, 0),   # close East   ~5 km
    (10, 10, 0),   # medium NE    ~14 km
    ( 0, 12, 0),   # medium North ~12 km
    ( 8,  3, 0),   # ESE          ~9 km
]

print(f'Running {len(hit_points)} simulations...')
runs1 = run_grid(base_grid, hit_points)
plot_grid(runs1, title='Grid — 4 hit points, no angle constraints')
export_grid_excel(runs1, prefix='grid1')


# In[21]:


# Cross-product: 4 hit points × 2 angle constraints = 8 IACPN runs
hit_points2 = [
    ( 20,  0, 0),
    (20, 20, 0)
]
hit_angles2 = [(-20, None), (-45, None)]

runs2 = run_grid(base_grid, hit_points2, hit_angles=hit_angles2,
                 hit_angle_range=10)
plot_grid(runs2, title='IACPN — 4 hit points × 2 angle constraints')
export_grid_excel(runs2, prefix='grid2')


# ## 16 — Trajectory optimisation (direct collocation, CasADi + IPOPT)
# 
# `collocation_solve()` in `sim/collocate.py` finds the optimal **(a_nV(t), a_nH(t))**
# schedule using **Hermite-Simpson direct collocation** with **IPOPT** as the solver.
# 
# **Why CasADi + IPOPT instead of SLSQP?**
# 
# The SLSQP-based `optimize_trajectory()` in `sim/opt_trajectory.py` failed due to three compounding issues:
# 1. **Vanishing gradient** — smooth RangeOnLine sigmoid ≈ 0 when angles are far from target → SLSQP had no signal.
# 2. **Model-reality mismatch** — Euler at dt_coll=0.6 s ≠ fine sim at dt=0.2 s → "solved" collocation gave 700 m miss in replay.
# 3. **Scale** — 641 variables, 483 equality constraints with finite-diff Jacobians → SLSQP hit iteration limit every run.
# 
# `collocate.py` fixes all three:
# - **CasADi automatic differentiation** — exact analytic Jacobians and Hessians.
# - **IPOPT interior-point** — designed for large-scale NLP; handles hundreds of nodes reliably.
# - **Hermite-Simpson** — 4th-order accurate; collocation model matches the continuous dynamics closely.
# - **Hit-line formulation** — terminal landing point can lie anywhere on the approach line; angle tolerance enforced as a hard constraint.
# 
# **Three bugs fixed in `collocate.py` (were preventing convergence):**
# 1. Warm start used angle-constrained IACPN → chaotic trajectory → bad initial guess.  Fixed: no-angle IACPN warm start.
# 2. No bounds on γV/γH at intermediate nodes → IPOPT drove angles to ±1 000°.  Fixed: added γV ∈ [−π/2, π/2] bounds.
# 3. `dgV = g·cos(γV)/v` singular at v=0 → constraint violation ≈ 10²³ → "Restoration Failed".  Fixed: clamp v to max(v, grav_turn_v_min) in angle-rate denominators.
# 
# **Merit function:**
# 
#     J = w_ctrl      · ControlEffortMerit(pct_mean)
#       + w_miss      · MissMerit(miss_m²)
#       + w_time      · FlightTimeMerit(T_f)
#       + w_final_line· FinalLineMerit(FL_smooth)
# 
# - **ControlEffortMerit** — softplus penalty on mean control usage (% of limit). Dead zone to 50%, ~100 at 80%, steep above 95%.
# - **MissMerit** — squared miss distance m². Hit-line mode: `w_miss` defaults to 0.0 (position free on approach line). Point-target mode: `w_miss` defaults to 1.0.
# - **FlightTimeMerit** — total flight time T_f.
# - **FinalLineMerit** — quartic-ReLU penalty on short FinalLine approach segment: `50·max(0, D − FL_smooth)⁴`. Active only in hit-line mode; set `w_final_line=0` to disable.
# 
# **Terminal constraints (hit-line mode, hard — not in objective):**
# - Terminal position on the hit line through the target in the desired approach direction
# - |γV_f − γV_des| ≤ angle_tol_deg
# - cos(γH_f − γH_des) ≥ cos(angle_tol_deg)  (handles ±180° wrap)

# In[ ]:


from sim.collocate import collocation_solve, plot_collocated
from sim.physics import Params

# Demo params — vehicle that reliably reaches the 10 km target
p_opt = Params(
    m_pay=50, m_prop=495, m_str=50, isp=260, thrust_kn=31, burn_max=20,
    cd=0.4, cd_fall=1.2, cl=0.0, cd_ctrl=0.3, diam=0.5,
    launch_angle=45, launch_azimuth=45,
    x_target=10, y_target=10, z_target=0, a_lat_max=3.0,
    hit_gamma_v=-45.0,   # desired impact flight-path angle (deg)
    hit_gamma_h=45.0,    # desired impact azimuth (deg)
)

result_opt = collocation_solve(
    p_opt,
    N1=15,              # Hermite-Simpson segments, powered phase
    N2=30,              # Hermite-Simpson segments, coast phase
    angle_tol_deg=5.0,  # hard angle tolerance (deg)
    w_ctrl=1.0,         # ControlEffortMerit weight
    w_miss=0.0,         # MissMerit weight (0 = position free on hit line)
    w_time=1e-3,        # FlightTimeMerit weight
    w_final_line=0.1,   # FinalLineMerit weight — penalises short approach alignment
    ipopt_opts={'ipopt.print_level': 0, 'print_time': False},
)

print()
print("=== Optimisation summary ===")
for k, v in result_opt['summary'].items():
    print(f"  {k:<35} {v}")


# In[23]:


plot_collocated(result_opt, title='Hermite-Simpson collocation — hit-line merit')


# ### 16b — Tighter angle tolerance
# 
# Reducing `angle_tol_deg` forces the optimizer to hold the approach angle
# more precisely.  The hit-line offset increases as the tighter angle constraint
# limits which approach trajectories are feasible.
# 

# In[ ]:


from dataclasses import replace

# Steeper approach angle, tighter tolerance
p_tight = replace(
    p_opt,
    hit_gamma_v=-30.0,
    hit_gamma_h=45.0,
)

result_tight = collocation_solve(
    p_tight,
    N1=15,
    N2=30,
    angle_tol_deg=3.0,   # tighter — harder to achieve long on-line segment
    w_ctrl=1.0,
    w_time=1e-3,
    ipopt_opts={'ipopt.print_level': 0, 'print_time': False},
)

print()
print("=== Tight-angle summary ===")
for k, v in result_tight['summary'].items():
    print(f"  {k:<35} {v}")


# In[ ]:


plot_collocated(result_tight, title='Hermite-Simpson collocation — tighter tolerance')


# ## 17 — Optimizer grid run
# 
# `run_grid(..., optimise=True)` runs `optimize_trajectory()` for every
# cell in the grid (Cartesian product of hit_points × hit_angles).
# 
# **Key differences from IACPN grid:**
# - Each run uses the direct-collocation optimizer (single stage, RangeOnLine merit).
# - `solved = True` when `optimised_miss_m < 1 m`.
# - Summary includes `range_on_line_km` — approach segment length.
# - Bar chart: green = solved, red = not solved.
# - `opt_kwargs` passes optimizer settings to every run
#   (N, a_max_g, w_range, w_effort, w_time, angle_tol_deg, max_iter).
# 

# In[ ]:


# Optimizer grid — two hit points × two angle constraints (all with γV and γH)
opt_points = [
    (-15, 15, 0),   # NW target ~21 km
    ( 0, -15, 0),   # S  target ~15 km
]
opt_angles = [
    (-45,  45.0),   # steep dive, approach from SW heading NE
    (-45, -135.0),  # steep dive, approach from NE heading SW
]

runs_opt = run_grid(
    base_grid,
    opt_points,
    hit_angles=opt_angles,
    optimise=True,
    opt_kwargs=dict(
        N=60,
        a_max_g=20.0,
        w_range=1.0,
        w_effort=1e-3,
        w_time=1e-4,
        angle_tol_deg=5.0,
        max_iter=300,
    ),
)
plot_grid(runs_opt, title="Optimizer grid — all angle-constrained (γV + γH)")


# ### 17b — Optimizer grid with tighter angle tolerance
# 
# Using `angle_tol_deg=3.0` forces more precise approach angles. All runs include both γV and γH constraints.

# In[ ]:


# Optimizer grid with tighter angle tolerance — all runs angle-constrained (γV + γH)
opt_points_b = [
    (-20, -20, 0),  # SW target ~28 km
    (  0, -20, 0),  # S  target ~20 km
]
opt_angles_b = [
    (-45,  45.0),   # steep dive, approach from SW heading NE
    (-30, -90.0),   # shallower dive, approach from E heading W
]

runs_opt_b = run_grid(
    base_grid,
    opt_points_b,
    hit_angles=opt_angles_b,
    optimise=True,
    opt_kwargs=dict(
        N=60,
        a_max_g=3.0,
        w_range=1.0,
        w_effort=1e-3,
        w_time=1e-4,
        angle_tol_deg=3.0,
        max_iter=300,
    ),
)
plot_grid(runs_opt_b, title="Optimizer grid — tighter tolerance, all angle-constrained")

